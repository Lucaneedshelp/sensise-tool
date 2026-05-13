import base64
import hashlib
import json
import os
import re
import sys
import urllib.error
import urllib.request
from pathlib import Path

from openpyxl import load_workbook
from pypdf import PdfReader


ROOT = Path(__file__).resolve().parents[1]
SOURCE_DIR = ROOT / "source-knowledge"
BUILD_DIR = ROOT / "build" / "knowledge"
OUTPUT_JSON = BUILD_DIR / "azure-documents.json"
API_VERSION = "2024-07-01"
MAX_CONTENT_CHARS = 3600
CHUNK_CHARS = 2800


LANG_PRIORITY = {
    "de": 0,
    "en": 1,
    "fr": 2,
    "pl": 3,
    "unknown": 4,
}


def load_env():
    env_path = ROOT / ".env"
    if not env_path.exists():
        return
    for raw_line in env_path.read_text(encoding="utf-8", errors="ignore").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ.setdefault(key.strip(), value.strip().strip("\"'"))


def detect_language(path):
    name = path.name.lower()
    if "datenblatt_de" in name or "_de." in name or "_de_" in name:
        return "de"
    if "datasheet_en" in name or "manual_en" in name or "guide_en" in name or "_en." in name or "_en_" in name:
        return "en"
    if "fiche" in name or "_fr." in name or "_fr_" in name:
        return "fr"
    if "_pl." in name or "_pl_" in name:
        return "pl"
    return "unknown"


def normalize_product(path):
    name = path.stem
    name = re.sub(r"(?i)(_datenblatt_de|_datasheet_en|_datasheet_fr|_datasheet_pl|_fiche_technique_fr|_fiche_technique|_manual_en|_quick_start_guide_en)$", "", name)
    name = re.sub(r"(?i)(_payload_decoder|_downlink_encoder|_downlink_information_en)$", "", name)
    name = name.replace("_", " ")
    name = re.sub(r"\s+", " ", name).strip()
    return name


def infer_type(path):
    parts = [part.lower() for part in path.parts]
    if "preise" in parts:
        return "preise"
    if "produktlisten" in parts:
        return "produkte"
    if "schnittstellen" in parts:
        return "schnittstellen"
    if "interne-regeln" in parts:
        return "antwortregel"
    if "testfragen" in parts:
        return "testfragen"
    if "datenblaetter" in parts:
        return "datenblatt"
    return "wissen"


def clean_text(text):
    text = text.replace("\x00", " ")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    text = re.sub(r"(?im)^\s*(page|seite)\s+\d+\s*(of|von)?\s*\d*\s*$", "", text)
    return text.strip()


def read_pdf(path):
    try:
        reader = PdfReader(str(path))
        pages = []
        for page in reader.pages:
            pages.append(page.extract_text() or "")
        return clean_text("\n\n".join(pages))
    except Exception as exc:
        print(f"Warnung: PDF konnte nicht gelesen werden: {path.name}: {exc}", file=sys.stderr)
        return ""


def read_xlsx(path):
    try:
        workbook = load_workbook(path, data_only=True, read_only=True)
        lines = []
        for sheet in workbook.worksheets:
            lines.append(f"# Tabelle: {sheet.title}")
            for row in sheet.iter_rows(values_only=True):
                values = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
                if values:
                    lines.append(" | ".join(values))
        return clean_text("\n".join(lines))
    except Exception as exc:
        print(f"Warnung: Excel konnte nicht gelesen werden: {path.name}: {exc}", file=sys.stderr)
        return ""


def read_text(path):
    return clean_text(path.read_text(encoding="utf-8", errors="ignore"))


def file_text(path):
    suffix = path.suffix.lower()
    if suffix == ".pdf":
        return read_pdf(path)
    if suffix == ".xlsx":
        return read_xlsx(path)
    if suffix in {".txt", ".md", ".csv"}:
        return read_text(path)
    return ""


def compact_text(text):
    lines = []
    seen = set()
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line:
            if lines and lines[-1]:
                lines.append("")
            continue
        normalized = re.sub(r"\W+", "", line.lower())
        if len(normalized) > 18 and normalized in seen:
            continue
        seen.add(normalized)
        lines.append(line)
    return clean_text("\n".join(lines))


def chunk_text(text, max_chars=MAX_CONTENT_CHARS):
    if len(text) <= max_chars:
        return [text]

    chunks = []
    current = []
    current_len = 0
    for block in re.split(r"\n\s*\n", text):
        block = block.strip()
        if not block:
            continue
        if current and current_len + len(block) + 2 > CHUNK_CHARS:
            chunks.append("\n\n".join(current))
            current = []
            current_len = 0
        if len(block) > CHUNK_CHARS:
            for start in range(0, len(block), CHUNK_CHARS):
                part = block[start:start + CHUNK_CHARS].strip()
                if part:
                    chunks.append(part)
            continue
        current.append(block)
        current_len += len(block) + 2
    if current:
        chunks.append("\n\n".join(current))
    return chunks


def content_fingerprint(text):
    simplified = re.sub(r"\W+", "", text.lower())
    return hashlib.sha256(simplified.encode("utf-8")).hexdigest()


def doc_id(value):
    return base64.urlsafe_b64encode(value.encode("utf-8")).decode("ascii").rstrip("=")


def aliases_for(product, path):
    words = {product}
    compact = product.replace(" ", "")
    if compact:
        words.add(compact)
    for token in re.split(r"[\s_()\-+]+", product):
        if len(token) >= 3:
            words.add(token)
    name = path.name.lower()
    if "x-logic" in name or "pc-lr" in name:
        words.update(["X-Logic", "X Logic", "X Logix", "PC-LR-1", "Wasserzaehler", "Wasserzähler"])
    if "sab07" in name:
        words.update(["SAB07", "SAB 07", "Ventilposition", "Ventilstellung", "Motorposition", "Motorstellung"])
    return ", ".join(sorted(words))


def should_skip_language_duplicate(path, grouped_paths):
    language = detect_language(path)
    if language == "de":
        return False

    product = normalize_product(path).lower()
    same_product = grouped_paths.get(product, [])
    has_german = any(detect_language(candidate) == "de" for candidate in same_product)
    is_datasheet = infer_type(path) == "datenblatt"
    return is_datasheet and has_german


def build_documents():
    files = [path for path in SOURCE_DIR.rglob("*") if path.is_file()]
    grouped_paths = {}
    for path in files:
        grouped_paths.setdefault(normalize_product(path).lower(), []).append(path)

    documents = []
    seen_fingerprints = set()
    skipped_language = 0
    skipped_duplicate = 0
    skipped_empty = 0

    for path in sorted(files):
        if should_skip_language_duplicate(path, grouped_paths):
            skipped_language += 1
            continue

        text = compact_text(file_text(path))
        if not text:
            skipped_empty += 1
            continue

        product = normalize_product(path)
        source_type = infer_type(path)
        language = detect_language(path)
        alias_text = aliases_for(product, path)
        header = [
            f"Produkt/Quelle: {product}",
            f"Typ: {source_type}",
            f"Sprache: {language}",
            f"Synonyme/Aliase: {alias_text}",
            f"Datei: {path.relative_to(ROOT)}",
        ]
        enriched_text = "\n".join(header) + "\n\n" + text

        for index, chunk in enumerate(chunk_text(enriched_text), start=1):
            fingerprint = content_fingerprint(chunk[:2600])
            if fingerprint in seen_fingerprints:
                skipped_duplicate += 1
                continue
            seen_fingerprints.add(fingerprint)

            title = product if index == 1 else f"{product} Teil {index}"
            documents.append({
                "id": doc_id(f"{path.relative_to(ROOT)}::{index}"),
                "title": title[:250],
                "content": chunk[:MAX_CONTENT_CHARS],
                "source": str(path.relative_to(ROOT)).replace("\\", "/"),
                "type": source_type,
                "product": product[:250],
            })

    return documents, {
        "inputFiles": len(files),
        "documents": len(documents),
        "skippedLanguageDuplicates": skipped_language,
        "skippedExactDuplicates": skipped_duplicate,
        "skippedEmpty": skipped_empty,
        "contentBytes": sum(len(doc["content"].encode("utf-8")) for doc in documents),
    }


def request_json(url, api_key, body):
    request = urllib.request.Request(
        url,
        data=json.dumps(body, ensure_ascii=False).encode("utf-8"),
        headers={
            "Content-Type": "application/json",
            "api-key": api_key,
        },
        method="POST",
    )
    with urllib.request.urlopen(request, timeout=120) as response:
        return json.loads(response.read().decode("utf-8"))


def upload_documents(documents):
    endpoint = os.environ.get("AZURE_SEARCH_ENDPOINT", "").rstrip("/")
    index_name = os.environ.get("AZURE_SEARCH_INDEX", "")
    api_key = os.environ.get("AZURE_SEARCH_API_KEY", "")
    if not endpoint or not index_name or not api_key:
        raise RuntimeError("AZURE_SEARCH_ENDPOINT, AZURE_SEARCH_INDEX oder AZURE_SEARCH_API_KEY fehlt.")

    url = f"{endpoint}/indexes/{index_name}/docs/index?api-version={API_VERSION}"
    total = 0
    for start in range(0, len(documents), 500):
        batch = [{"@search.action": "mergeOrUpload", **doc} for doc in documents[start:start + 500]]
        result = request_json(url, api_key, {"value": batch})
        failed = [
            item for item in result.get("value", [])
            if not item.get("succeeded", item.get("status", False))
        ]
        if failed:
            raise RuntimeError(f"Azure Upload Fehler: {failed[:3]}")
        total += len(batch)
    return total


def main():
    load_env()
    BUILD_DIR.mkdir(parents=True, exist_ok=True)
    documents, stats = build_documents()
    OUTPUT_JSON.write_text(json.dumps(documents, ensure_ascii=False, indent=2), encoding="utf-8")

    print(json.dumps(stats, ensure_ascii=False, indent=2))
    print(f"Ausgabe: {OUTPUT_JSON.relative_to(ROOT)}")

    if "--upload" in sys.argv:
        uploaded = upload_documents(documents)
        print(f"Azure Upload abgeschlossen: {uploaded} Dokumente")


if __name__ == "__main__":
    try:
        main()
    except urllib.error.HTTPError as exc:
        body = exc.read().decode("utf-8", errors="ignore")
        print(f"Azure HTTP Fehler ({exc.code}): {body}", file=sys.stderr)
        sys.exit(1)
    except Exception as exc:
        print(f"Fehler: {exc}", file=sys.stderr)
        sys.exit(1)
